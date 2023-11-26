from tkinter import *
import openpyxl
from vari_1 import *
path = "full.xlsx"
wb_obj = openpyxl.load_workbook(path)
cell = []
button=[]
button_text=[]
main_word = []
count=0  
imp=0

def check(s2, s1):
    if type(s2) == str and type(s1) == str:
        if (s2.upper().count(s1) > 0):
            return True
        else:
            return False
    else:
        return False
result=0
ppl=0

class Elder:
    color = ["Aqua","Cyan", "Red", "Yellow", "Lime", "Orange", "Pink"]
    sheet_name=wb_obj.sheetnames
    sheet_name.sort()
    color.sort()
    row=0
    col=0
    sheet_obj=0
    imp_button=[]

    def __init__(self, master):

        myFrame5 = Frame(master)
        myFrame5.pack(side=TOP) 

        # both are using the same
        myFrame1 = Frame(master)
        myFrame1.pack(side=TOP)

        myFrame12 = Frame(master)
        myFrame12.pack(side=TOP)


        myFrame2 = Frame(master)
        myFrame2.pack(side=TOP)

        myFrame34 = Frame(master,width=1000)
        myFrame34.pack(side=TOP)

        myFrame3 = Frame(myFrame34,width=100, padx=1, pady=15)
        myFrame3.pack(side=TOP)

        myFrame4 = Frame(myFrame34,width=100, padx=1, pady=20)
        myFrame4.pack(side=TOP)

        self.cbh =0
        # Create Button
        # for i in range(65,86):
        #     if i < 73:
        #         inside=myFrame1
        #         clr="Pink"
        #     else:
        #         inside = myFrame2
        #         clr="Lime"
        #     i=chr(i)
        #     self.Button = Button(inside, text=i,command=lambda m=i: self.fillup_alphabet(m))
        #     self.Button.configure(bg=clr)
        #     self.Button.pack(side="left")


        for idx, num in enumerate(Elder.sheet_name):
            if idx <= 12:
                frame_change = myFrame1
            else:
                frame_change = myFrame12
            
            self.Button = Button(frame_change, text=num,command=lambda m=num: self.root_word(m))

            self.Button.configure(text=num.title())
            a = ord(num.split(" ")[-1][0].lower())%len(Elder.color)
            for j in range(len(Elder.color) + 1):
                if a == j:
                    self.Button.configure(bg=Elder.color[j])
                    break
            self.Button.pack(side="left")

        # Create text widget and specify size.
        self.Text = Text(myFrame2, height=1.46, width=16)
        self.Text.pack(side=LEFT)

        # Create Button
        self.Button1 = Button(myFrame2, text="Search", bg='lime',command=self.Search)
        self.Button1.pack(side=LEFT)

        # create Button
        self.Button2 = Button(myFrame2, text="Clear Alphabet",bg='lime', command=self.Clear_alphabet)
        self.Button2.pack(side=LEFT)

        # Create Button
        self.Button1 = Button(myFrame2, text="A+", bg='lime',command=self.add)
        self.Button1.pack(side=LEFT)

        # Create Scrollbar
        self.text3 = Text(myFrame3, wrap="none",height=8,width=75)
        self.vsb3 = Scrollbar(myFrame3,orient="vertical", command=self.text3.yview,)
        self.text3.configure(yscrollcommand=self.vsb3.set)
        self.vsb3.pack(side="right", fill="y")
        self.text3.pack(fill="both", expand=True)

        # Create Scrollbar
        self.text = Text(myFrame4, wrap="none",height=8,width=75)
        self.vsb = Scrollbar(myFrame4,orient="vertical", command=self.text.yview,)
        self.text.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.text.pack(fill="both", expand=True)

        # OUTPUT BOX
        self.output = Text(myFrame5,  height=18, width=80)
        self.output.grid(row=0,column=2,columnspan=2)

        # Label
        self.label1=Label(myFrame2, text=result)
        self.label1.pack(side=RIGHT)

    def root_word(self,text):
        global result
        dict1={
               "python":python_word,"django":django_word,"api":django_rest_api_word,
               "database":database_word,"cloud":cloud_word,"intro":intro_word,"program":program_word,
               "html": html_word,"project": project_word,"api testing":api_testing_word,
               "dts": dts_word,"os": os_word,"aws": aws_word,
               "azure": azure_word,
               "meter": meter_word,
               "datascience": dc_word,
               "german": german_word,
               "pytest": pytest_word,
               "azure": azure_word,
               "os": os_word,
               "rohan": rohan_word,
               "selenium": selenium_word,
               "vgs": vgs_word,
               }

        self.text3.delete(1.0, END)
        # self.text.delete(1.0, END)
        Elder.sheet_obj = wb_obj.active
        Elder.row = Elder.sheet_obj.max_row
        Elder.col = Elder.sheet_obj.max_column

        # print(f"before button press -->  {Elder.sheet_obj.title}")
        for i, s in enumerate(wb_obj.sheetnames):
            if s == text:
                wb_obj.active = i
                imp=dict1[wb_obj.sheetnames[i]]
                # print(text,s)
                break
        # print(f"After button press -->  {Elder.sheet_obj.title}")
        # print("#"*20)
        imp = sorted([i.title() for i in imp])
        length = 0
        clp=0
        imp_count = 0
        for i in imp:
            cbc = ""
            # if length == 0:
            #     cbc = ""
            # if length>75 or len(i)>55:
            if length + len(str(i)) >= 85:
                cbc = "\n"
                length = 0
            length += len(str(i))
            self.text3.insert("end", cbc)
            self.btn = Button(text=i,fg="black", padx=1, pady=1,command=lambda m=i : self.fillup(m))
            clp += 1
            a = ord(i[0].lower())%len(Elder.color)
            for j in range(len(Elder.color) + 1):
                if a == j:
                    self.btn.configure(bg=Elder.color[j])
                    break
            self.text3.pack(side="right")
            self.text3.window_create("end", window=self.btn)
            imp_count += 1
        result = Elder.sheet_obj.title.split(" ")[-1] +" active "+str(imp_count)
        # print(Elder.sheet_obj.title)
        self.label1.configure(text=result)
        # height total =35
        self.output.configure(height=17)
        a=imp_count//5 *0.8
        self.text3.configure(height=a)
        self.text.configure(height=16 - a)
        self.Search()
        self.text.delete(1.0, END)
        self.output.delete(1.0, END)
        if Elder.sheet_obj.title != text:
            self.root_word(text)

    def add(self):
        self.cbh+=1
        if self.cbh%2==0:
            self.output.configure(height=17)
            self.text.configure(height=8)
            self.text3.configure(height=8)
        else:
            self.output.configure(height=27)
            self.text.configure(height=3)
            self.text3.configure(height=3)

    def delete(self):
        self.text.delete(1.0, END)
        button.clear()
        button_text.clear()
        cell.clear()
        self.text.delete(1.0, END)

    def Clear_alphabet(self):
        self.Text.delete(1.0, END)

    def fillup(self,text):
        self.Text.delete(1.0, END)
        self.Text.insert(END,text)
        self.Search()
        # print(text)

    def second_fillup(self, text):
        self.Text.insert(END, " ")
        self.Text.insert(END, text)
        self.Search()

    def fillup_alphabet(self,text):
        self.Text.insert(END,text)
        self.Search()

    def delete2(self):
        for i in main_word:
            try:
                i.destroy()
            except:
                print("{} is missing".format(i))
        main_word.clear()

    def Search(self):
        global count
        count+=1
        if count==1:
            self.delete()
            count=0
        for i in range(1, wb_obj.active.max_row + 1):
            cell_obj = wb_obj.active.cell(row=i, column=1)
            s2 = cell_obj.value
            s1 = self.Text.get("1.0", "end-1c").upper()
            s3 = Elder.sheet_obj.cell(row=i, column=2).value
            if check(s2, s1) or check(s2,s1):
                cell.append(s3)
        length=0
        clp=0
        last_word=[1,2]
        # print("#" * 24)
        for i in cell:
            if i not in button_text and isinstance(i,str):
                cbc = " "
                if length == 0:
                    cbc = ""
                if length+len(str(i.replace(" ","").replace("\n","")))>85:
                    cbc="\n"
                    length=0
                self.text.insert("end", cbc)
                self.btn = Button(text=i,fg="black",command=lambda m=i : self.inner_search(m))
                if type(i)==str:
                    self.btn.configure(text=i.title())
                self.text.pack(side="right")
                self.text.window_create("end", window=self.btn)
                length += len(str(i.replace(" ","").replace("\n","")))
                clp+=1
                last_word.append(len(str(i)))
                a = clp%(len(Elder.color))
                for j in range(len(Elder.color) + 1):
                    if a == j:
                        self.btn.configure(bg=Elder.color[j])
                        break
                button.append(self.btn)
                button_text.append(i)

    def inner_search(self,Text):
        self.output.delete(1.0, END)
        a = Text
        for i in range(1, Elder.row + 1):
            cell_obj = Elder.sheet_obj.cell(row=i, column=2)
            s2 = cell_obj.value
            s1 = a
            if s2 == s1:
                self.output.insert(END, "Question:- ")
                # self.output.insert(END, "\nQuestion:- ")
                cell_obj = Elder.sheet_obj.cell(row=i, column=2)
                self.output.insert(END, cell_obj.value)
                self.output.insert(END, "\n\n")
                cell_obj = Elder.sheet_obj.cell(row=i, column=3)
                self.output.insert(END, cell_obj.value)
                self.output.insert(END,"\n")
                self.output.insert(END, "-"*80)
                self.output.insert(END, "\n\n")


root = Tk()
root.title("Search Field")
root.geometry('650x680') # wXh
e = Elder(root)
root.mainloop()


